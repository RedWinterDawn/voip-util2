#!/usr/bin/python2.6

import openpyxl
import psycopg2
import sys
import rateupdate

CONFIGURATOR = '10.125.252.170'
CDR = 'cdr'

#variables
datas = False
count = 0
pushes = 1
fileName = "/var/www/uploads/onvoy-ratedeck.xlsx"
connection = "dbname='ratedeck' user='postgres' host='%s' " %(CDR)

#load file
wb = openpyxl.load_workbook(fileName, use_iterators = True)
wsname =  wb.get_sheet_names()
ws = wb.get_sheet_by_name(name = wsname[0]) 
date = ws['B3'].value

##connect to db
try:
    db = psycopg2.connect(connection)
    cur = db.cursor()
    print 'connected to db'
except:
    print 'failed to connect to db'
    sys.exit(1)

##update validto
query = "UPDATE onvoy_domestic SET validto = '%s' where validto is NULL" %(date)
try:
    cur.execute(query)
    db.commit()
    print 'update success'
except:
    print 'failed to update'
    sys.exit(1)

for row in ws.iter_rows():
    if row[0].value == 'XXXXXX':
        print 'inter: %s      intra: %s' %(row[1].value, row[2].value)
    if datas == True and row[0].value !='XXXXXX':
        if count == 0:
            query = "INSERT INTO onvoy_domestic (lrn, inter, intra, validfrom) VALUES (%s, %s, %s, '%s')" %(row[0].value, row[1].value, row[2].value, date)
            count = count + 1
        elif count < 10000:
            query = "%s, (%s, %s, %s, '%s')" %(query, row[0].value, row[1].value, row[2].value, date)
            count  = count + 1
        else:    
            try:
                cur.execute(query)
                db.commit()
                print pushes, ' ',
                pushes = pushes +1
                count = 0
                query = ''
            except psycopg2.Error as e:
                print e.pgerror        
                pass
    if row[0].value == 'LRN':
        datas = True
try:
    cur.execute(query)
    db.commit()
except psycopg2.Error as e:
    print e.pgerror
    pass
db.close()

print ''
rateupdate.update('Onvoy', 'onvoy_domestic', 'lrn')
print ''
print 'Completed'
